from openpyxl import load_workbook
from get_from_divvy import get_from_divvy
from get_from_jobber import get_from_jobber
import os

wb = load_workbook(f"{os.getenv('WORKBOOK_PATH')}/workbook.xlsx")

"""
HOW TO USE
Set the environment variables:
    WORKBOOK_PATH
    DIVVY_API_KEY
    JOBBER_API_KEY
    
In your workbook, any jobber auto pages should be named
    J_A_OBJECT_OBJECT_OBJECT
    D_A_OBJECT_OBJECT_OBJECT
where the object is your target object, and the sub objects are its children
    Divvy objects should be named like "objects" pluralized and lowercased
    Jobber objects should be named like "Object" capitalized and single
Your headers should be 
    property_property_property
where each property is a property of the object

Save this into workbook.xlsx
Your output will be in update_workbook.xlsx

"""




for sheet_name in wb.sheetnames:

    sheet = wb[sheet_name]

    if sheet_name.startswith('J_A_'):
        get_from_jobber(sheet, sheet_name)
    elif sheet_name.startswith('D_A_'):
        get_from_divvy(sheet, sheet_name)
    else:
        continue

# Save the modified workbook
wb.save(f"{os.getenv('WORKBOOK_PATH')}/update_workbook.xlsx")