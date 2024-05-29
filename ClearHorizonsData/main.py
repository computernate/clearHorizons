import requests
from openpyxl import load_workbook

# Load the spreadsheet
wb = load_workbook('workbook.xlsx')

# Iterate over all sheets
for sheet_name in wb.sheetnames:
    # Check if sheet name starts with 'JOBBER_AUTO_'
    if sheet_name.startswith('JOBBER_AUTO_'):
        sheet = wb[sheet_name]

        # Read the headers from the first row
        headers = [cell.value for cell in sheet[1]]

        # Extract XXX from the sheet name
        table_name = sheet_name[len('JOBBER_AUTO_'):]

        # Construct the GraphQL query based on the headers and table name
        query_fields = []
        for header in headers:
            nested_fields = header.split('_')
            nested_structure = ''
            for field in nested_fields[:-1]:
                nested_structure += f'{field}{{'
            nested_structure += nested_fields[-1]
            nested_structure += '}' * (len(nested_fields) - 1)
            query_fields.append(nested_structure)

        query = f'''
        query {table_name.capitalize()}Query {{
          {table_name} {{
            nodes {{
              {",".join(query_fields)}
            }}
          }}
        }}
        '''
        print(query)
        # Send the query to Jobber's API
        url = "https://api.getjobber.com/api/graphql"
        api_key = "eyJhbGciOiJIUzI1NiJ9.eyJzdWIiOjIxNTQ4NzQsImlzcyI6Imh0dHBzOi8vYXBpLmdldGpvYmJlci5jb20iLCJjbGllbnRfaWQiOiI2MDdlNjRjMi00OTdkLTQyOWEtODJhYi01MTVjODgxM2M2NWQiLCJzY29wZSI6InJlYWRfY2xpZW50cyB3cml0ZV9jbGllbnRzIHJlYWRfcmVxdWVzdHMgd3JpdGVfcmVxdWVzdHMgcmVhZF9xdW90ZXMgd3JpdGVfcXVvdGVzIHJlYWRfam9icyB3cml0ZV9qb2JzIHJlYWRfc2NoZWR1bGVkX2l0ZW1zIHdyaXRlX3NjaGVkdWxlZF9pdGVtcyByZWFkX2ludm9pY2VzIHdyaXRlX2ludm9pY2VzIHJlYWRfdXNlcnMgd3JpdGVfdXNlcnMgcmVhZF9jdXN0b21fZmllbGRfY29uZmlndXJhdGlvbnMgd3JpdGVfY3VzdG9tX2ZpZWxkX2NvbmZpZ3VyYXRpb25zIHJlYWRfdGltZV9zaGVldHMiLCJhcHBfaWQiOiI2MDdlNjRjMi00OTdkLTQyOWEtODJhYi01MTVjODgxM2M2NWQiLCJ1c2VyX2lkIjoyMTU0ODc0LCJhY2NvdW50X2lkIjoxMTQ1NzkwLCJleHAiOjE3MTY3MDE4OTB9.j5xMDwIBQG4IxTNMj4Aw5OqWQG2wxZrfYV2c7gtjeTE"

        # Set the headers
        request_headers = {
            "Authorization": f"Bearer {api_key}",
            "X-JOBBER-GRAPHQL-VERSION": "2023-11-15"
        }

        response = requests.post(url, headers=request_headers, json={'query':query})
        # Check if the request was successful
        if response.status_code == 200:
            data = response.json()
            # Process the data as needed
        else:
            print(f"Failed to fetch data from Jobber's API {response.status_code}")
            data=None

        # Parse the response and populate the sheet
        for row_num, node in enumerate(data['data'][table_name]['nodes'], start=2):
            for col_num, field in enumerate(headers, start=1):
                nested_fields = field.split('_')
                nested_value = node
                for nested_field in nested_fields:
                    nested_value = nested_value.get(nested_field, {})
                sheet.cell(row=row_num, column=col_num, value=nested_value)

# Save the updated spreadsheet
wb.save('updated_workbook.xlsx')